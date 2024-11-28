import pandas as pd
import os.path
import pyodbc
import datetime
from openpyxl import load_workbook
from tinydb import TinyDB
from openpyxl.utils.dataframe import dataframe_to_rows
from send_mail import Send_Mail
from io import BytesIO
from plyer import notification
from Parts.db_connect import db_con
from Past_Backup import backup
from dotenv import load_dotenv
try:
    load_dotenv(dotenv_path='Parts/private_data.env')

    PAST_SQL =os.getenv('FARK_RP_SQL')

    file = os.path.join(os.getcwd(), r"Rp_Modules\Reports", "Past_Backup.xlsx")
    Report_File_Path = os.path.join(os.getcwd(), r"Rp_Modules\Reports", "Sinif_Liste_Fark_Rapor.xlsx")
    if os.path.isfile(file) == True:# Program İlk Çalıştığında geçmiş bir yedek yoksa geçmiş yedeği alır ve kapanır
        todays_date = datetime.date.today()
        day = todays_date.isoweekday()
        xlsx = pd.ExcelFile(file) # geçmiş excel dosyasını pandasa tanımladık
        past_pd = pd.read_excel(xlsx,index_col=None) # geçmiş exceli dataframe çevirip index kolonunu kaldırdık.
        new_pd =pd.read_sql(PAST_SQL,db_con.conn,index_col=None) #güncel verileri sorgu ile çekip dataframe aktardık

        col1 = os.getenv('FARK_IMP_COL_1')
        col2 = os.getenv('FARK_IMP_COL_2')
        col3 = os.getenv('FARK_IMP_COL_3')
        Uniq_Col = os.getenv('FARK_UNQ_COL_4')
        filter_col = os.getenv('FARK_FILTER_COL')

        changing_data = pd.DataFrame(columns=past_pd.columns) 
        past_pd[[col1,col2,col3]] = past_pd[[col1,col2,col3]].astype({col1:int,col2:int,col3:int})
        new_pd[[col1,col2,col3]] = new_pd[[col1,col2,col3]].astype({col1:int,col2:int,col3:int})

        non_existent_data = past_pd[~past_pd[Uniq_Col].isin(new_pd[Uniq_Col])]
        added_data = new_pd[~new_pd[Uniq_Col].isin(past_pd[Uniq_Col])]

        merge_pd = pd.merge(past_pd, new_pd, on=Uniq_Col, how='outer', indicator=True)
        merge_pd = merge_pd[merge_pd['_merge'] == 'both']
        merge_pd = merge_pd[
                                    (merge_pd[col2+'_x'] != merge_pd[col2+'_y']) |
                                    (merge_pd[col1+'_x'] != merge_pd[col1+'_y']) |
                                    (merge_pd[col3+'_x'] != merge_pd[col3+'_y'])
                                ]                           
        
        for i in range(len(merge_pd)):
            item=merge_pd.iloc[i]
            changing_data = pd.concat([changing_data, past_pd[(past_pd[Uniq_Col] == item[Uniq_Col])]], ignore_index = True)
            changing_data = pd.concat([changing_data, new_pd[(new_pd[Uniq_Col] == item[Uniq_Col])]], ignore_index = True)
        today = datetime.datetime.today().strftime('%d-%m-%Y')
        non_existent_data['Rapor_Tarihi'] = today
        added_data['Rapor_Tarihi'] = today
        changing_data['Rapor_Tarihi'] = today
        
        if os.path.isfile(Report_File_Path) == True:
            if day == 1:
                writer = pd.ExcelWriter(Report_File_Path, engine ='xlsxwriter',engine_kwargs={'options':{'strings_to_urls': False}})
                non_existent_data.to_excel(writer,index=False,sheet_name="Çıkan Ürünler")
                changing_data.to_excel(writer,index=False,sheet_name="Değişen Ürünler")
                added_data.to_excel(writer,index=False,sheet_name="Yeni Ürünler")
                writer.close() 
            else:
                book = load_workbook(Report_File_Path)

                sheet = book["Çıkan Ürünler"]
                start_row = sheet.max_row
                for row in dataframe_to_rows(non_existent_data, index=False, header=False):
                    sheet.append(row)
                
                sheet = book["Değişen Ürünler"]
                start_row = sheet.max_row
                for row in dataframe_to_rows(changing_data, index=False, header=False):
                    sheet.append(row)

                sheet = book["Yeni Ürünler"]
                start_row = sheet.max_row
                for row in dataframe_to_rows(added_data, index=False, header=False):
                    sheet.append(row)

                book.save(Report_File_Path)
        else:
            writer = pd.ExcelWriter(Report_File_Path, engine ='xlsxwriter',engine_kwargs={'options':{'strings_to_urls': False}})
            non_existent_data.to_excel(writer,index=False,sheet_name="Çıkan Ürünler")
            changing_data.to_excel(writer,index=False,sheet_name="Değişen Ürünler")
            added_data.to_excel(writer,index=False,sheet_name="Yeni Ürünler")
            writer.close() 
        db_path = os.path.join(os.getcwd(), "Rp_Modules", "reports.json")
        db = TinyDB(db_path)    
        table_object = db.table('sinif_liste_fark')    
        df_cikan = pd.read_excel(Report_File_Path,sheet_name="Çıkan Ürünler") 
        df_degisen = pd.read_excel(Report_File_Path,sheet_name="Değişen Ürünler") 
        df_eklenen = pd.read_excel(Report_File_Path,sheet_name="Yeni Ürünler") 
        for item in table_object.all():
            item_rp_time = item["rp_send_days"]
            if day in item_rp_time:
                opt=item['school_filter']
                exists_opt_cikan=[]
                exists_opt_degisen=[]
                exists_opt_eklenen=[]
                for x in opt:
                    if x in df_cikan[filter_col].values:
                        exists_opt_cikan.append(x)
                    if x in df_degisen[filter_col].values:
                        exists_opt_degisen.append(x)
                    if x in df_eklenen[filter_col].values:
                        exists_opt_eklenen.append(x)
                df_cikan_excel = df_cikan[df_cikan[filter_col].isin(exists_opt_cikan)]
                df_degisen_excel = df_degisen[df_degisen[filter_col].isin(exists_opt_degisen)]
                df_eklenen_excel = df_eklenen[df_eklenen[filter_col].isin(exists_opt_eklenen)]
                col_index = item['col_filter']
                col_index.append(35) 
                col_index.sort()
                df_cikan_excel = df_cikan_excel.iloc[:,col_index]
                df_degisen_excel = df_degisen_excel.iloc[:,col_index]
                df_eklenen_excel = df_eklenen_excel.iloc[:,col_index]
                excel_buffer = BytesIO()

                with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                    df_cikan_excel.to_excel(writer, index=False, sheet_name="Çıkan Ürünler")
                    df_degisen_excel.to_excel(writer, index=False, sheet_name="Değişen Ürünler")
                    df_eklenen_excel.to_excel(writer, index=False, sheet_name="Yeni Ürünler")
                excel_buffer.seek(0)
                Send_Mail(item["Subject"],item['Mail'],item['Full_Name'],excel_buffer)
                del excel_buffer
    else:
        x = backup()
        x.new_backup() 
    
except Exception as e:
    print(f'Error: {e}')
    notification.notify(
        title='Rapor HATA Bildirimi',
        message=f'Sınıf Listelerindeki Fark Raporu {e} hatasından dolayı hazırlanamadı ve ilgili kişilere mail gönderilemedi.',
        app_name='OS Oto Rapor Gönderimi',
    )
